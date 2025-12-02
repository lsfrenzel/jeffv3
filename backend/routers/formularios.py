from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import List, Optional
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from backend.database import get_db
from backend.models.formularios import Formulario, Pergunta, OpcaoResposta, FormularioEnvio, Resposta
from backend.models.empresas import Empresa
from backend.models.usuarios import Usuario
from backend.schemas import formularios as schemas
from backend.auth.security import obter_usuario_atual as get_current_user

router = APIRouter(prefix="/api/formularios", tags=["formularios"])

@router.get("/", response_model=List[schemas.FormularioResumo])
async def listar_formularios(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Lista todos os formulários com resumo de envios e respostas"""
    formularios = db.query(Formulario).options(
        joinedload(Formulario.perguntas),
        joinedload(Formulario.envios)
    ).all()
    
    resultado = []
    for form in formularios:
        total_respostas = db.query(FormularioEnvio).filter(
            FormularioEnvio.formulario_id == form.id,
            FormularioEnvio.respondido == True
        ).count()
        
        resultado.append(schemas.FormularioResumo(
            id=form.id,
            titulo=form.titulo,
            descricao=form.descricao,
            tipo=form.tipo,
            ativo=form.ativo,
            data_criacao=form.data_criacao,
            total_perguntas=len(form.perguntas),
            total_envios=len(form.envios),
            total_respostas=total_respostas
        ))
    
    return resultado

@router.get("/{formulario_id}", response_model=schemas.Formulario)
async def obter_formulario(
    formulario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Obtém um formulário completo com todas as perguntas e opções"""
    formulario = db.query(Formulario).options(
        joinedload(Formulario.perguntas).joinedload(Pergunta.opcoes)
    ).filter(Formulario.id == formulario_id).first()
    
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    return formulario

@router.post("/", response_model=schemas.Formulario)
async def criar_formulario(
    formulario_data: schemas.FormularioCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Cria um novo formulário personalizado"""
    formulario = Formulario(
        titulo=formulario_data.titulo,
        descricao=formulario_data.descricao,
        tipo=formulario_data.tipo,
        ativo=formulario_data.ativo,
        criado_por=current_user.id
    )
    db.add(formulario)
    db.flush()
    
    for ordem, perg_data in enumerate(formulario_data.perguntas):
        pergunta = Pergunta(
            formulario_id=formulario.id,
            texto=perg_data.texto,
            tipo=perg_data.tipo,
            categoria=perg_data.categoria,
            descricao_categoria=perg_data.descricao_categoria,
            ordem=ordem,
            obrigatoria=perg_data.obrigatoria
        )
        db.add(pergunta)
        db.flush()
        
        for opc_data in perg_data.opcoes:
            opcao = OpcaoResposta(
                pergunta_id=pergunta.id,
                texto=opc_data.texto,
                descricao=opc_data.descricao,
                valor=opc_data.valor
            )
            db.add(opcao)
    
    db.commit()
    db.refresh(formulario)
    
    return db.query(Formulario).options(
        joinedload(Formulario.perguntas).joinedload(Pergunta.opcoes)
    ).filter(Formulario.id == formulario.id).first()

@router.put("/{formulario_id}", response_model=schemas.Formulario)
async def atualizar_formulario(
    formulario_id: int,
    formulario_data: schemas.FormularioUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Atualiza um formulário existente"""
    formulario = db.query(Formulario).filter(Formulario.id == formulario_id).first()
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    if formulario_data.titulo is not None:
        formulario.titulo = formulario_data.titulo
    if formulario_data.descricao is not None:
        formulario.descricao = formulario_data.descricao
    if formulario_data.ativo is not None:
        formulario.ativo = formulario_data.ativo
    
    db.commit()
    db.refresh(formulario)
    
    return db.query(Formulario).options(
        joinedload(Formulario.perguntas).joinedload(Pergunta.opcoes)
    ).filter(Formulario.id == formulario.id).first()

@router.delete("/{formulario_id}")
async def excluir_formulario(
    formulario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Exclui um formulário"""
    formulario = db.query(Formulario).filter(Formulario.id == formulario_id).first()
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    db.delete(formulario)
    db.commit()
    
    return {"message": "Formulário excluído com sucesso"}

@router.post("/{formulario_id}/perguntas", response_model=schemas.Pergunta)
async def adicionar_pergunta(
    formulario_id: int,
    pergunta_data: schemas.PerguntaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Adiciona uma pergunta a um formulário"""
    formulario = db.query(Formulario).filter(Formulario.id == formulario_id).first()
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    max_ordem = db.query(func.max(Pergunta.ordem)).filter(
        Pergunta.formulario_id == formulario_id
    ).scalar() or 0
    
    pergunta = Pergunta(
        formulario_id=formulario_id,
        texto=pergunta_data.texto,
        tipo=pergunta_data.tipo,
        categoria=pergunta_data.categoria,
        descricao_categoria=pergunta_data.descricao_categoria,
        ordem=max_ordem + 1,
        obrigatoria=pergunta_data.obrigatoria
    )
    db.add(pergunta)
    db.flush()
    
    for opc_data in pergunta_data.opcoes:
        opcao = OpcaoResposta(
            pergunta_id=pergunta.id,
            texto=opc_data.texto,
            descricao=opc_data.descricao,
            valor=opc_data.valor
        )
        db.add(opcao)
    
    db.commit()
    db.refresh(pergunta)
    
    return db.query(Pergunta).options(
        joinedload(Pergunta.opcoes)
    ).filter(Pergunta.id == pergunta.id).first()

@router.delete("/perguntas/{pergunta_id}")
async def excluir_pergunta(
    pergunta_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Exclui uma pergunta"""
    pergunta = db.query(Pergunta).filter(Pergunta.id == pergunta_id).first()
    if not pergunta:
        raise HTTPException(status_code=404, detail="Pergunta não encontrada")
    
    db.delete(pergunta)
    db.commit()
    
    return {"message": "Pergunta excluída com sucesso"}

@router.post("/enviar", response_model=schemas.FormularioEnvio)
async def enviar_formulario(
    envio_data: schemas.FormularioEnvioCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Cria um link único para envio do formulário"""
    formulario = db.query(Formulario).filter(Formulario.id == envio_data.formulario_id).first()
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    if not formulario.ativo:
        raise HTTPException(status_code=400, detail="Formulário está inativo")
    
    envio = FormularioEnvio(
        formulario_id=envio_data.formulario_id,
        empresa_id=envio_data.empresa_id,
        email_destinatario=envio_data.email_destinatario,
        nome_destinatario=envio_data.nome_destinatario,
        enviado_por=current_user.id
    )
    db.add(envio)
    db.commit()
    db.refresh(envio)
    
    return envio

@router.get("/envios/", response_model=List[schemas.FormularioEnvioDetalhado])
async def listar_envios(
    formulario_id: Optional[int] = None,
    respondido: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Lista todos os envios de formulários"""
    query = db.query(FormularioEnvio).options(
        joinedload(FormularioEnvio.formulario),
        joinedload(FormularioEnvio.empresa)
    )
    
    if formulario_id:
        query = query.filter(FormularioEnvio.formulario_id == formulario_id)
    if respondido is not None:
        query = query.filter(FormularioEnvio.respondido == respondido)
    
    envios = query.order_by(FormularioEnvio.data_envio.desc()).all()
    
    resultado = []
    for envio in envios:
        resultado.append(schemas.FormularioEnvioDetalhado(
            id=envio.id,
            formulario_id=envio.formulario_id,
            empresa_id=envio.empresa_id,
            codigo_unico=envio.codigo_unico,
            email_destinatario=envio.email_destinatario,
            nome_destinatario=envio.nome_destinatario,
            respondido=envio.respondido,
            data_envio=envio.data_envio,
            data_resposta=envio.data_resposta,
            formulario=envio.formulario,
            empresa_nome=envio.empresa.razao_social if envio.empresa else None
        ))
    
    return resultado

@router.get("/envios/{envio_id}/respostas", response_model=schemas.EnvioComRespostas)
async def obter_respostas_envio(
    envio_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Obtém as respostas de um envio específico"""
    envio = db.query(FormularioEnvio).options(
        joinedload(FormularioEnvio.formulario),
        joinedload(FormularioEnvio.empresa),
        joinedload(FormularioEnvio.respostas).joinedload(Resposta.pergunta)
    ).filter(FormularioEnvio.id == envio_id).first()
    
    if not envio:
        raise HTTPException(status_code=404, detail="Envio não encontrado")
    
    respostas_vis = []
    for resp in envio.respostas:
        opcao_texto = None
        if resp.valor_numerico is not None:
            opcao = db.query(OpcaoResposta).filter(
                OpcaoResposta.pergunta_id == resp.pergunta_id,
                OpcaoResposta.valor == resp.valor_numerico
            ).first()
            if opcao:
                opcao_texto = opcao.texto
        
        respostas_vis.append(schemas.RespostaVisualizacao(
            pergunta_texto=resp.pergunta.texto,
            categoria=resp.pergunta.categoria,
            valor_numerico=resp.valor_numerico,
            valor_texto=resp.valor_texto,
            opcao_texto=opcao_texto
        ))
    
    return schemas.EnvioComRespostas(
        id=envio.id,
        formulario_id=envio.formulario_id,
        empresa_id=envio.empresa_id,
        codigo_unico=envio.codigo_unico,
        email_destinatario=envio.email_destinatario,
        nome_destinatario=envio.nome_destinatario,
        respondido=envio.respondido,
        data_envio=envio.data_envio,
        data_resposta=envio.data_resposta,
        respostas=respostas_vis,
        formulario_titulo=envio.formulario.titulo if envio.formulario else None,
        empresa_nome=envio.empresa.razao_social if envio.empresa else None
    )

@router.get("/{formulario_id}/estatisticas", response_model=schemas.EstatisticasFormulario)
async def obter_estatisticas(
    formulario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Obtém estatísticas de um formulário"""
    formulario = db.query(Formulario).filter(Formulario.id == formulario_id).first()
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    total_envios = db.query(FormularioEnvio).filter(
        FormularioEnvio.formulario_id == formulario_id
    ).count()
    
    total_respostas = db.query(FormularioEnvio).filter(
        FormularioEnvio.formulario_id == formulario_id,
        FormularioEnvio.respondido == True
    ).count()
    
    taxa_resposta = (total_respostas / total_envios * 100) if total_envios > 0 else 0
    
    media_por_pergunta = {}
    perguntas = db.query(Pergunta).filter(Pergunta.formulario_id == formulario_id).all()
    for pergunta in perguntas:
        if pergunta.tipo == "escala":
            media = db.query(func.avg(Resposta.valor_numerico)).filter(
                Resposta.pergunta_id == pergunta.id,
                Resposta.valor_numerico.isnot(None)
            ).scalar()
            if media is not None:
                media_por_pergunta[pergunta.id] = {
                    "texto": pergunta.texto[:50] + "..." if len(pergunta.texto) > 50 else pergunta.texto,
                    "media": round(float(media), 2)
                }
    
    return schemas.EstatisticasFormulario(
        formulario_id=formulario_id,
        titulo=formulario.titulo,
        total_envios=total_envios,
        total_respostas=total_respostas,
        taxa_resposta=round(taxa_resposta, 1),
        media_por_pergunta=media_por_pergunta
    )

@router.get("/{formulario_id}/exportar-excel")
async def exportar_estatisticas_excel(
    formulario_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Exporta as estatísticas e respostas de um formulário para Excel"""
    formulario = db.query(Formulario).options(
        joinedload(Formulario.perguntas).joinedload(Pergunta.opcoes)
    ).filter(Formulario.id == formulario_id).first()
    
    if not formulario:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    envios = db.query(FormularioEnvio).options(
        joinedload(FormularioEnvio.empresa),
        joinedload(FormularioEnvio.respostas)
    ).filter(
        FormularioEnvio.formulario_id == formulario_id,
        FormularioEnvio.respondido == True
    ).all()
    
    wb = Workbook()
    
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_resumo['A1'] = "Relatório de Estatísticas"
    ws_resumo['A1'].font = Font(bold=True, size=16)
    ws_resumo.merge_cells('A1:D1')
    
    ws_resumo['A3'] = "Formulário:"
    ws_resumo['B3'] = formulario.titulo
    ws_resumo['A4'] = "Data de Exportação:"
    ws_resumo['B4'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    total_envios = db.query(FormularioEnvio).filter(
        FormularioEnvio.formulario_id == formulario_id
    ).count()
    total_respostas = len(envios)
    taxa = (total_respostas / total_envios * 100) if total_envios > 0 else 0
    
    ws_resumo['A6'] = "Total de Envios:"
    ws_resumo['B6'] = total_envios
    ws_resumo['A7'] = "Total de Respostas:"
    ws_resumo['B7'] = total_respostas
    ws_resumo['A8'] = "Taxa de Resposta:"
    ws_resumo['B8'] = f"{round(taxa, 1)}%"
    
    ws_resumo['A10'] = "Média por Pergunta"
    ws_resumo['A10'].font = Font(bold=True, size=12)
    
    row = 11
    headers = ["Pergunta", "Categoria", "Média", "Qtd Respostas"]
    for col, header in enumerate(headers, 1):
        cell = ws_resumo.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    perguntas = sorted(formulario.perguntas, key=lambda p: p.ordem)
    for pergunta in perguntas:
        if pergunta.tipo == "escala":
            media = db.query(func.avg(Resposta.valor_numerico)).filter(
                Resposta.pergunta_id == pergunta.id,
                Resposta.valor_numerico.isnot(None)
            ).scalar()
            qtd = db.query(Resposta).filter(
                Resposta.pergunta_id == pergunta.id,
                Resposta.valor_numerico.isnot(None)
            ).count()
            
            ws_resumo.cell(row=row, column=1, value=pergunta.texto).border = border
            ws_resumo.cell(row=row, column=2, value=pergunta.categoria or "-").border = border
            ws_resumo.cell(row=row, column=3, value=round(float(media), 2) if media else 0).border = border
            ws_resumo.cell(row=row, column=4, value=qtd).border = border
            row += 1
    
    for col in range(1, 5):
        ws_resumo.column_dimensions[get_column_letter(col)].width = 40 if col == 1 else 20
    
    ws_respostas = wb.create_sheet("Respostas Detalhadas")
    
    headers = ["Destinatário", "Email", "Empresa", "Data Resposta"]
    for pergunta in perguntas:
        headers.append(pergunta.texto[:50] + "..." if len(pergunta.texto) > 50 else pergunta.texto)
    
    for col, header in enumerate(headers, 1):
        cell = ws_respostas.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row = 2
    for envio in envios:
        ws_respostas.cell(row=row, column=1, value=envio.nome_destinatario or "Anônimo").border = border
        ws_respostas.cell(row=row, column=2, value=envio.email_destinatario or "-").border = border
        ws_respostas.cell(row=row, column=3, value=envio.empresa.razao_social if envio.empresa else "-").border = border
        ws_respostas.cell(row=row, column=4, value=envio.data_resposta.strftime("%d/%m/%Y %H:%M") if envio.data_resposta else "-").border = border
        
        respostas_dict = {r.pergunta_id: r for r in envio.respostas}
        col = 5
        for pergunta in perguntas:
            resposta = respostas_dict.get(pergunta.id)
            if resposta:
                if resposta.valor_numerico is not None:
                    valor = resposta.valor_numerico
                elif resposta.valor_texto:
                    valor = resposta.valor_texto
                else:
                    valor = "-"
            else:
                valor = "-"
            ws_respostas.cell(row=row, column=col, value=valor).border = border
            col += 1
        row += 1
    
    for col in range(1, len(headers) + 1):
        ws_respostas.column_dimensions[get_column_letter(col)].width = 20
    
    ws_respostas.row_dimensions[1].height = 40
    
    ws_dist = wb.create_sheet("Distribuição Respostas")
    
    row = 1
    for pergunta in perguntas:
        if pergunta.tipo == "escala":
            ws_dist.cell(row=row, column=1, value=pergunta.texto).font = Font(bold=True)
            ws_dist.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            
            headers = ["Valor", "Descrição", "Quantidade", "Percentual"]
            for col, header in enumerate(headers, 1):
                cell = ws_dist.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            row += 1
            
            total_respostas_pergunta = db.query(Resposta).filter(
                Resposta.pergunta_id == pergunta.id,
                Resposta.valor_numerico.isnot(None)
            ).count()
            
            for opcao in sorted(pergunta.opcoes, key=lambda o: o.valor or 0):
                count = db.query(Resposta).filter(
                    Resposta.pergunta_id == pergunta.id,
                    Resposta.valor_numerico == opcao.valor
                ).count()
                pct = (count / total_respostas_pergunta * 100) if total_respostas_pergunta > 0 else 0
                
                ws_dist.cell(row=row, column=1, value=opcao.valor).border = border
                ws_dist.cell(row=row, column=2, value=opcao.texto).border = border
                ws_dist.cell(row=row, column=3, value=count).border = border
                ws_dist.cell(row=row, column=4, value=f"{round(pct, 1)}%").border = border
                row += 1
            
            row += 1
    
    for col in [1, 2, 3, 4]:
        ws_dist.column_dimensions[get_column_letter(col)].width = 25 if col == 2 else 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"estatisticas_{formulario.titulo.replace(' ', '_')[:30]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/seed-lideranca")
async def criar_formulario_lideranca(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Cria o formulário padrão de Avaliação das Lideranças"""
    existente = db.query(Formulario).filter(
        Formulario.titulo == "Avaliação das Lideranças"
    ).first()
    
    if existente:
        return {"message": "Formulário já existe", "id": existente.id}
    
    opcoes_escala = [
        {"texto": "1 - Nunca", "descricao": "O comportamento ou resultado nunca é observado.", "valor": 1},
        {"texto": "2 - Raramente", "descricao": "O comportamento ou resultado é observado em pouquíssimas ocasiões.", "valor": 2},
        {"texto": "3 - Ocasionalmente", "descricao": "Ocorre, mas não é consistente ou não faz parte da rotina.", "valor": 3},
        {"texto": "4 - Frequente", "descricao": "O comportamento ou resultado é uma prática comum do líder.", "valor": 4},
        {"texto": "5 - Muito frequente", "descricao": "O comportamento ou resultado é uma prática constante e bem estabelecida.", "valor": 5},
    ]
    
    perguntas_lideranca = [
        {
            "categoria": "Liderança, Pessoas e Desenvolvimento",
            "descricao_categoria": "Foco na capacidade do líder de gerir, inspirar e desenvolver sua equipe.",
            "perguntas": [
                "1. O líder define e comunica claramente os papéis e as responsabilidades de cada membro da equipe.",
                "2. O líder acompanha e cobra regularmente o atingimento das metas individuais e do time, mantendo o foco em resultados.",
                "3. O líder mantém o foco profissional, sem deixar que as relações pessoais comprometam a autoridade e as decisões.",
                "4. O líder fornece feedback formal e informal à equipe sobre seu desempenho e pontos de melhoria de forma estruturada.",
                "5. O líder apoia e orienta o desenvolvimento individual dos colaboradores (ex: PDI), buscando o crescimento de suas competências.",
                "6. O líder identifica as necessidades de capacitação da equipe e propõe/busca treinamentos e recursos de desenvolvimento.",
                "7. O líder atua no nível hierárquico esperado (tático/estratégico), delegando tarefas operacionais à equipe quando apropriado.",
            ]
        },
        {
            "categoria": "Comunicação, Clima e Cultura",
            "descricao_categoria": "Foco na transparência, resolução de conflitos e manutenção de um ambiente produtivo.",
            "perguntas": [
                "8. O líder garante que a comunicação seja clara e estruturada na área, assegurando que todos estejam alinhados com as prioridades.",
                "9. O líder aborda e resolve conflitos internos de forma estruturada, justa e imediata, buscando o aprendizado.",
                "10. O líder estimula e incentiva a equipe a propor e implementar melhorias e inovações nos processos e rotinas.",
                "11. O líder reconhece e valoriza publicamente os bons desempenhos e os esforços da equipe, motivando os colaboradores.",
                "12. O líder realiza momentos de \"escuta ativa\" (one-on-one, reuniões) para entender o clima e os desafios individuais/coletivos.",
                "13. As ações do líder demonstram que ele prioriza a satisfação do cliente (interno ou externo) e age com base em seus feedbacks.",
            ]
        },
        {
            "categoria": "Processos, Qualidade e Resultados",
            "descricao_categoria": "Foco na disciplina, na utilização de dados e na gestão da rotina.",
            "perguntas": [
                "14. O líder assegura que os processos sob sua responsabilidade sejam seguidos de forma padronizada e com disciplina.",
                "15. O líder monitora de perto os indicadores (KPIs) da área e utiliza esses dados para orientar a gestão no dia a dia.",
                "16. O líder baseia suas decisões em dados e fatos concretos, evitando percepções subjetivas (\"achismo\").",
                "17. O líder utiliza metodologias estruturadas (como PDCA ou 5W2H) para a análise e resolução de problemas recorrentes.",
                "18. O líder conduz reuniões produtivas, com pauta clara, tempo definido e registro/acompanhamento de planos de ação.",
                "19. O líder utiliza ferramentas de Gestão à Vista (quadros, dashboards) para manter a transparência e o foco nos resultados da área.",
            ]
        }
    ]
    
    formulario = Formulario(
        titulo="Avaliação das Lideranças",
        descricao="Questionário para avaliação de competências e comportamentos de liderança, com escala de 1 a 5.",
        tipo="padrao",
        ativo=True,
        criado_por=current_user.id
    )
    db.add(formulario)
    db.flush()
    
    ordem = 0
    for grupo in perguntas_lideranca:
        for texto_pergunta in grupo["perguntas"]:
            pergunta = Pergunta(
                formulario_id=formulario.id,
                texto=texto_pergunta,
                tipo="escala",
                categoria=grupo["categoria"],
                descricao_categoria=grupo["descricao_categoria"],
                ordem=ordem,
                obrigatoria=True
            )
            db.add(pergunta)
            db.flush()
            
            for opc in opcoes_escala:
                opcao = OpcaoResposta(
                    pergunta_id=pergunta.id,
                    texto=opc["texto"],
                    descricao=opc["descricao"],
                    valor=opc["valor"]
                )
                db.add(opcao)
            
            ordem += 1
    
    db.commit()
    
    return {"message": "Formulário de Avaliação das Lideranças criado com sucesso", "id": formulario.id}


router_public = APIRouter(prefix="/formulario", tags=["formulario_publico"])

@router_public.get("/responder/{codigo}")
async def pagina_responder_formulario(
    codigo: str,
    request: Request,
    db: Session = Depends(get_db)
):
    """Página pública para responder um formulário"""
    from fastapi.templating import Jinja2Templates
    templates = Jinja2Templates(directory="templates")
    
    envio = db.query(FormularioEnvio).options(
        joinedload(FormularioEnvio.formulario).joinedload(Formulario.perguntas).joinedload(Pergunta.opcoes)
    ).filter(FormularioEnvio.codigo_unico == codigo).first()
    
    if not envio:
        return templates.TemplateResponse("formulario_erro.html", {
            "request": request,
            "erro": "Formulário não encontrado"
        })
    
    if envio.respondido:
        return templates.TemplateResponse("formulario_respondido.html", {
            "request": request,
            "formulario": envio.formulario
        })
    
    if not envio.formulario.ativo:
        return templates.TemplateResponse("formulario_erro.html", {
            "request": request,
            "erro": "Este formulário não está mais disponível"
        })
    
    return templates.TemplateResponse("formulario_responder.html", {
        "request": request,
        "envio": envio,
        "formulario": envio.formulario,
        "perguntas": sorted(envio.formulario.perguntas, key=lambda p: p.ordem)
    })

@router_public.post("/responder/{codigo}")
async def submeter_respostas(
    codigo: str,
    respostas_data: schemas.FormularioRespostaSubmit,
    db: Session = Depends(get_db)
):
    """Submete as respostas de um formulário"""
    envio = db.query(FormularioEnvio).options(
        joinedload(FormularioEnvio.formulario).joinedload(Formulario.perguntas)
    ).filter(FormularioEnvio.codigo_unico == codigo).first()
    
    if not envio:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    if envio.respondido:
        raise HTTPException(status_code=400, detail="Este formulário já foi respondido")
    
    if not envio.formulario.ativo:
        raise HTTPException(status_code=400, detail="Este formulário não está mais disponível")
    
    for resp_data in respostas_data.respostas:
        resposta = Resposta(
            envio_id=envio.id,
            pergunta_id=resp_data.pergunta_id,
            valor_numerico=resp_data.valor_numerico,
            valor_texto=resp_data.valor_texto,
            valores_multiplos=resp_data.valores_multiplos
        )
        db.add(resposta)
    
    envio.respondido = True
    envio.data_resposta = datetime.now()
    
    db.commit()
    
    return {"message": "Respostas enviadas com sucesso!"}

@router_public.get("/api/{codigo}")
async def obter_formulario_publico(
    codigo: str,
    db: Session = Depends(get_db)
):
    """Obtém dados do formulário para preenchimento público"""
    envio = db.query(FormularioEnvio).options(
        joinedload(FormularioEnvio.formulario).joinedload(Formulario.perguntas).joinedload(Pergunta.opcoes)
    ).filter(FormularioEnvio.codigo_unico == codigo).first()
    
    if not envio:
        raise HTTPException(status_code=404, detail="Formulário não encontrado")
    
    if envio.respondido:
        raise HTTPException(status_code=400, detail="Este formulário já foi respondido")
    
    if not envio.formulario.ativo:
        raise HTTPException(status_code=400, detail="Este formulário não está mais disponível")
    
    perguntas_ordenadas = sorted(envio.formulario.perguntas, key=lambda p: p.ordem)
    
    return {
        "envio_id": envio.id,
        "formulario": {
            "titulo": envio.formulario.titulo,
            "descricao": envio.formulario.descricao,
            "perguntas": [
                {
                    "id": p.id,
                    "texto": p.texto,
                    "tipo": p.tipo,
                    "categoria": p.categoria,
                    "descricao_categoria": p.descricao_categoria,
                    "obrigatoria": p.obrigatoria,
                    "opcoes": [
                        {
                            "id": o.id,
                            "texto": o.texto,
                            "descricao": o.descricao,
                            "valor": o.valor
                        } for o in sorted(p.opcoes, key=lambda x: x.valor or 0)
                    ]
                } for p in perguntas_ordenadas
            ]
        }
    }

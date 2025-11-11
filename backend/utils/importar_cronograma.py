import openpyxl
from sqlalchemy.orm import Session
from backend.database import SessionLocal
from backend.models import CronogramaProjeto, Usuario, Empresa
from backend.models.cronograma import StatusProjeto
from datetime import datetime
import re


def normalizar_cnpj(cnpj):
    if not cnpj:
        return None
    cnpj_str = str(cnpj).strip()
    cnpj_limpo = re.sub(r'[^\d]', '', cnpj_str)
    if len(cnpj_limpo) == 14:
        return f"{cnpj_limpo[:2]}.{cnpj_limpo[2:5]}.{cnpj_limpo[5:8]}/{cnpj_limpo[8:12]}-{cnpj_limpo[12:]}"
    return cnpj_str


def importar_cronograma_planilha(caminho_planilha: str, db: Session = None):
    fecha_db = False
    if db is None:
        db = SessionLocal()
        fecha_db = True
    
    try:
        print(f"Abrindo planilha: {caminho_planilha}")
        wb = openpyxl.load_workbook(caminho_planilha, read_only=True, data_only=True)
        
        if 'CONSULTA INFO.' not in wb.sheetnames:
            print("ERRO: Aba 'CONSULTA INFO.' não encontrada")
            return
        
        ws = wb['CONSULTA INFO.']
        print(f"Processando {ws.max_row} linhas...")
        
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        print(f"Cabeçalhos encontrados: {headers[:10]}")
        
        mapa_colunas = {}
        for idx, header in enumerate(headers):
            mapa_colunas[header.upper()] = idx
        
        projetos_criados = 0
        projetos_atualizados = 0
        erros = 0
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                proposta = str(row[mapa_colunas.get('PROPOSTA', 0)]).strip() if row[mapa_colunas.get('PROPOSTA', 0)] else None
                if not proposta or proposta == 'None':
                    continue
                
                empresa_nome = str(row[mapa_colunas.get('EMPRESA', 1)]).strip() if row[mapa_colunas.get('EMPRESA', 1)] else None
                cnpj_raw = row[mapa_colunas.get('CNPJ', 2)]
                cnpj = normalizar_cnpj(cnpj_raw)
                sigla = str(row[mapa_colunas.get('SIGLA', 3)]).strip() if row[mapa_colunas.get('SIGLA', 3)] else None
                er = str(row[mapa_colunas.get('ER', 4)]).strip() if row[mapa_colunas.get('ER', 4)] else None
                porte = str(row[mapa_colunas.get('PORTE', 5)]).strip() if row[mapa_colunas.get('PORTE', 5)] else None
                solucao = str(row[mapa_colunas.get('SOLUÇÃO', 6)]).strip() if row[mapa_colunas.get('SOLUÇÃO', 6)] else None
                horas = float(row[mapa_colunas.get('HORAS', 7)]) if row[mapa_colunas.get('HORAS', 7)] else 0
                consultor_nome = str(row[mapa_colunas.get('CONSULTOR 1', 8)]).strip() if row[mapa_colunas.get('CONSULTOR 1', 8)] else None
                data_inicio = row[mapa_colunas.get('INÍCIO', 9)] if mapa_colunas.get('INÍCIO') is not None else None
                data_termino = row[mapa_colunas.get('TÉRMINO', 10)] if mapa_colunas.get('TÉRMINO') is not None else None
                endereco = str(row[mapa_colunas.get('ENDEREÇO', 11)]).strip() if row[mapa_colunas.get('ENDEREÇO', 11)] else None
                regiao = str(row[mapa_colunas.get('REGIÃO', 12)]).strip() if row[mapa_colunas.get('REGIÃO', 12)] else None
                municipio = str(row[mapa_colunas.get('MUNICÍPIO', 13)]).strip() if row[mapa_colunas.get('MUNICÍPIO', 13)] else None
                uf = str(row[mapa_colunas.get('UF', 14)]).strip() if row[mapa_colunas.get('UF', 14)] else None
                contato = str(row[mapa_colunas.get('CONTATO', 15)]).strip() if row[mapa_colunas.get('CONTATO', 15)] else None
                telefone = str(row[mapa_colunas.get('TELEFONE', 16)]).strip() if row[mapa_colunas.get('TELEFONE', 16)] else None
                celular = str(row[mapa_colunas.get('CELULAR', 17)]).strip() if row[mapa_colunas.get('CELULAR', 17)] else None
                
                if isinstance(data_inicio, datetime):
                    data_inicio = data_inicio.date()
                if isinstance(data_termino, datetime):
                    data_termino = data_termino.date()
                
                empresa_id = None
                if cnpj:
                    empresa = db.query(Empresa).filter(Empresa.cnpj == cnpj).first()
                    if empresa:
                        empresa_id = empresa.id
                
                consultor_id = None
                if consultor_nome and consultor_nome != 'None':
                    consultor = db.query(Usuario).filter(
                        Usuario.nome.ilike(f"%{consultor_nome}%")
                    ).first()
                    if consultor:
                        consultor_id = consultor.id
                
                projeto_existente = db.query(CronogramaProjeto).filter(
                    CronogramaProjeto.proposta == proposta
                ).first()
                
                if projeto_existente:
                    projeto_existente.empresa_id = empresa_id
                    projeto_existente.cnpj = cnpj
                    projeto_existente.sigla = sigla
                    projeto_existente.er = er
                    projeto_existente.porte = porte
                    projeto_existente.solucao = solucao
                    projeto_existente.horas_totais = horas
                    projeto_existente.consultor_id = consultor_id
                    projeto_existente.consultor_nome = consultor_nome
                    projeto_existente.data_inicio = data_inicio
                    projeto_existente.data_termino = data_termino
                    projeto_existente.endereco = endereco
                    projeto_existente.regiao = regiao
                    projeto_existente.municipio = municipio
                    projeto_existente.uf = uf
                    projeto_existente.contato = contato
                    projeto_existente.telefone = telefone
                    projeto_existente.celular = celular
                    projeto_existente.data_atualizacao = datetime.utcnow()
                    
                    if data_termino and data_termino < datetime.now().date():
                        projeto_existente.status = StatusProjeto.concluido
                    elif data_inicio and data_inicio <= datetime.now().date():
                        projeto_existente.status = StatusProjeto.em_andamento
                    
                    projetos_atualizados += 1
                else:
                    status = StatusProjeto.planejado
                    if data_termino and data_termino < datetime.now().date():
                        status = StatusProjeto.concluido
                    elif data_inicio and data_inicio <= datetime.now().date():
                        status = StatusProjeto.em_andamento
                    
                    novo_projeto = CronogramaProjeto(
                        proposta=proposta,
                        empresa_id=empresa_id,
                        cnpj=cnpj,
                        sigla=sigla,
                        er=er,
                        porte=porte,
                        solucao=solucao,
                        horas_totais=horas,
                        consultor_id=consultor_id,
                        consultor_nome=consultor_nome,
                        data_inicio=data_inicio,
                        data_termino=data_termino,
                        endereco=endereco,
                        regiao=regiao,
                        municipio=municipio,
                        uf=uf,
                        contato=contato,
                        telefone=telefone,
                        celular=celular,
                        status=status
                    )
                    db.add(novo_projeto)
                    projetos_criados += 1
                
                if (projetos_criados + projetos_atualizados) % 50 == 0:
                    db.commit()
                    print(f"Processadas {projetos_criados + projetos_atualizados} linhas...")
                
            except Exception as e:
                print(f"Erro na linha {i}: {str(e)}")
                erros += 1
                continue
        
        db.commit()
        
        print(f"\n=== Importação concluída ===")
        print(f"Projetos criados: {projetos_criados}")
        print(f"Projetos atualizados: {projetos_atualizados}")
        print(f"Erros: {erros}")
        print(f"Total processado: {projetos_criados + projetos_atualizados}")
        
        wb.close()
        
    except Exception as e:
        print(f"ERRO FATAL: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        if fecha_db:
            db.close()


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Uso: python importar_cronograma.py <caminho_planilha>")
        sys.exit(1)
    
    caminho = sys.argv[1]
    importar_cronograma_planilha(caminho)

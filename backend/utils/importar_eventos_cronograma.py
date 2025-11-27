#!/usr/bin/env python3
"""
Script para importar eventos do cronograma a partir da planilha Excel.
Formato dos eventos: [CATEGORIA]-[SIGLA_EMPRESA]
Categorias:
    C = Consultoria
    K = Kick-off
    F = Reuniao Final
    M = Mentoria
    T = T0 - Diagnostico
    P = Programado
    O = Outros
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import openpyxl
from datetime import datetime, date
from sqlalchemy.orm import Session
from backend.database import SessionLocal, engine, Base
from backend.models.cronograma import CronogramaEvento, CategoriaEvento, PeriodoEvento
from backend.models.usuarios import Usuario

CATEGORIA_MAP = {
    'C': CategoriaEvento.consultoria,
    'K': CategoriaEvento.kickoff,
    'F': CategoriaEvento.reuniao_final,
    'M': CategoriaEvento.mentoria,
    'T': CategoriaEvento.diagnostico,
    'P': CategoriaEvento.programado,
    'O': CategoriaEvento.outros,
}

def parse_evento_codigo(codigo: str):
    """Parse um codigo de evento no formato CATEGORIA-SIGLA"""
    if not codigo or not isinstance(codigo, str):
        return None, None
    
    codigo = codigo.strip()
    if '-' not in codigo:
        return None, None
    
    partes = codigo.split('-', 1)
    categoria_letra = partes[0].upper()
    sigla = partes[1] if len(partes) > 1 else None
    
    categoria = CATEGORIA_MAP.get(categoria_letra)
    
    return categoria, sigla


def importar_eventos(arquivo_xlsx: str, limpar_existentes: bool = False):
    """Importa eventos do arquivo Excel para o banco de dados"""
    
    if not os.path.exists(arquivo_xlsx):
        print(f"Erro: Arquivo {arquivo_xlsx} nao encontrado")
        return
    
    print(f"Abrindo arquivo: {arquivo_xlsx}")
    wb = openpyxl.load_workbook(arquivo_xlsx, data_only=True)
    
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    
    try:
        if limpar_existentes:
            print("Removendo eventos existentes...")
            db.query(CronogramaEvento).delete()
            db.commit()
        
        consultores = {}
        usuarios = db.query(Usuario).all()
        for u in usuarios:
            consultores[u.nome.lower()] = u.id
            partes_nome = u.nome.split()
            if len(partes_nome) >= 2:
                primeiro_ultimo = f"{partes_nome[0]} {partes_nome[-1]}".lower()
                consultores[primeiro_ultimo] = u.id
        
        print(f"Consultores encontrados no banco: {len(usuarios)}")
        
        if 'DADOS' in wb.sheetnames:
            ws = wb['DADOS']
            print("Processando planilha DADOS...")
            importar_da_planilha_dados(db, ws, consultores)
        elif 'CRONOGRAMA' in wb.sheetnames:
            ws = wb['CRONOGRAMA']
            print("Processando planilha CRONOGRAMA...")
            importar_da_planilha_cronograma(db, ws, consultores)
        else:
            print("Nenhuma planilha de cronograma encontrada")
            return
        
        db.commit()
        
        total = db.query(CronogramaEvento).count()
        print(f"\nImportacao concluida! Total de eventos: {total}")
        
    except Exception as e:
        print(f"Erro durante importacao: {e}")
        db.rollback()
        raise
    finally:
        db.close()
        wb.close()


def normalizar_nome(nome):
    """Normaliza um nome para comparacao"""
    if not nome:
        return ""
    nome = nome.lower().strip()
    nome = nome.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
    nome = nome.replace("é", "e").replace("è", "e").replace("ê", "e")
    nome = nome.replace("í", "i").replace("ì", "i").replace("î", "i")
    nome = nome.replace("ó", "o").replace("ò", "o").replace("õ", "o").replace("ô", "o")
    nome = nome.replace("ú", "u").replace("ù", "u").replace("û", "u")
    nome = nome.replace("ç", "c")
    return nome


def importar_da_planilha_dados(db: Session, ws, consultores: dict):
    """Importa eventos da planilha DADOS"""
    
    consultor_cols = {}
    for col in range(11, ws.max_column + 1, 2):
        nome_cell = ws.cell(row=2, column=col).value
        if nome_cell and isinstance(nome_cell, str):
            nome_planilha = normalizar_nome(nome_cell)
            consultor_id = None
            
            for nome_db, uid in consultores.items():
                nome_db_norm = normalizar_nome(nome_db)
                if nome_planilha == nome_db_norm:
                    consultor_id = uid
                    break
                partes_planilha = nome_planilha.split()
                partes_db = nome_db_norm.split()
                if len(partes_planilha) >= 2 and len(partes_db) >= 2:
                    if partes_planilha[0] == partes_db[0] and partes_planilha[-1] == partes_db[-1]:
                        consultor_id = uid
                        break
                    if partes_planilha[0] == partes_db[0]:
                        consultor_id = uid
                        break
            
            if consultor_id:
                consultor_cols[col] = {'id': consultor_id, 'nome': nome_cell, 'periodo': 'M'}
                consultor_cols[col + 1] = {'id': consultor_id, 'nome': nome_cell, 'periodo': 'T'}
                print(f"  Mapeado: {nome_cell} -> ID {consultor_id}")
    
    print(f"Colunas de consultores mapeadas: {len(consultor_cols)}")
    
    eventos_criados = 0
    
    for row in range(5, ws.max_row + 1):
        data_cell = ws.cell(row=row, column=2).value
        
        if not data_cell:
            continue
            
        if isinstance(data_cell, datetime):
            data_evento = data_cell.date()
        elif isinstance(data_cell, date):
            data_evento = data_cell
        else:
            continue
        
        if data_evento.year < 2020 or data_evento.year > 2030:
            continue
        
        for col, info in consultor_cols.items():
            cell_value = ws.cell(row=row, column=col).value
            
            if not cell_value or not isinstance(cell_value, str):
                continue
            
            cell_value = cell_value.strip()
            if not cell_value:
                continue
            
            categoria, sigla = parse_evento_codigo(cell_value)
            
            if not categoria:
                continue
            
            periodo = PeriodoEvento.manha if info['periodo'] == 'M' else PeriodoEvento.tarde
            
            evento = CronogramaEvento(
                data=data_evento,
                categoria=categoria,
                periodo=periodo,
                sigla_empresa=sigla,
                consultor_id=info['id'],
                titulo=cell_value
            )
            
            db.add(evento)
            eventos_criados += 1
            
            if eventos_criados % 100 == 0:
                print(f"  {eventos_criados} eventos processados...")
    
    print(f"Total de eventos criados: {eventos_criados}")


def importar_da_planilha_cronograma(db: Session, ws, consultores: dict):
    """Importa eventos da planilha CRONOGRAMA (layout alternativo)"""
    
    consultor_cols = {}
    header_row = 12
    
    for col in range(5, ws.max_column + 1):
        nome_cell = ws.cell(row=header_row - 1, column=col).value
        if nome_cell and isinstance(nome_cell, str) and nome_cell.strip():
            nome_lower = nome_cell.strip().lower()
            consultor_id = None
            for nome_db, uid in consultores.items():
                if nome_lower in nome_db or nome_db in nome_lower:
                    consultor_id = uid
                    break
            if consultor_id:
                consultor_cols[col] = {'id': consultor_id, 'nome': nome_cell}
    
    print(f"Colunas de consultores mapeadas: {len(consultor_cols)}")
    
    eventos_criados = 0
    
    for row in range(header_row + 1, ws.max_row + 1):
        data_cell = None
        for check_col in range(1, 5):
            cell_val = ws.cell(row=row, column=check_col).value
            if isinstance(cell_val, (datetime, date)):
                data_cell = cell_val
                break
        
        if not data_cell:
            continue
            
        if isinstance(data_cell, datetime):
            data_evento = data_cell.date()
        else:
            data_evento = data_cell
        
        if data_evento.year < 2020 or data_evento.year > 2030:
            continue
        
        for col, info in consultor_cols.items():
            cell_value = ws.cell(row=row, column=col).value
            
            if not cell_value or not isinstance(cell_value, str):
                continue
            
            cell_value = cell_value.strip()
            if not cell_value:
                continue
            
            categoria, sigla = parse_evento_codigo(cell_value)
            
            if not categoria:
                continue
            
            evento = CronogramaEvento(
                data=data_evento,
                categoria=categoria,
                periodo=PeriodoEvento.dia_todo,
                sigla_empresa=sigla,
                consultor_id=info['id'],
                titulo=cell_value
            )
            
            db.add(evento)
            eventos_criados += 1
    
    print(f"Total de eventos criados: {eventos_criados}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importar_eventos_cronograma.py <arquivo.xlsx> [--limpar]")
        print("  --limpar: Remove todos os eventos existentes antes de importar")
        sys.exit(1)
    
    arquivo = sys.argv[1]
    limpar = '--limpar' in sys.argv
    
    importar_eventos(arquivo, limpar)
